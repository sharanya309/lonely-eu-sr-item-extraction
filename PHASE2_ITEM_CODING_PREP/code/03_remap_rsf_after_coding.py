#!/usr/bin/env python3
"""
STEP 3 — RUN AFTER ITEM CODING. Re-map RSF domains back onto EVERY Phase-2 master_flat
item, so nothing is left uncoded.

For each master_flat item row (from item_disposition.csv):
  - disposition == new           -> take final_coding from the CODING OUTPUT (matched on item text)
  - disposition == already_coded -> use inherit_final_coding (from the Phase-1 32,444)
  - disposition == internal_dup  -> use the final_coding of its kept twin (maps_to_row_id)

Usage:
  python3 03_remap_rsf_after_coding.py --coded <coding_output.csv|.xlsx> [--text-col item_text_english] [--code-col final_coding]

The coding-output file must contain the coded NEW items with an item-text column and a
final domain column. Writes output/master_flat_with_rsf.csv (every item row + final_coding)
and reports how many rows were successfully assigned.
"""
import csv, os, sys, argparse
sys.path.insert(0, os.path.dirname(__file__))
from common_paths import MASTER_FLAT, OUT, norm, item_key
csv.field_size_limit(2**30)

ap = argparse.ArgumentParser()
ap.add_argument("--coded", required=True, help="coding-output file (new items + assigned domain)")
ap.add_argument("--text-col", default="item_text_english")
ap.add_argument("--code-col", default="final_coding")
a = ap.parse_args()

def load_any(path):
    if path.lower().endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True); ws = wb.active
        h = [c.value for c in next(ws.iter_rows(max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            yield dict(zip(h, row))
        wb.close()
    else:
        yield from csv.DictReader(open(path, encoding="utf-8", errors="replace"))

# coded NEW items: normalized text -> domain
coded_new = {}
for r in load_any(a.coded):
    k = norm(r.get(a.text_col))
    if k and r.get(a.code_col): coded_new.setdefault(k, r[a.code_col])
print(f"coded new-item domains loaded: {len(coded_new)}")

# master_flat rows by _mid
mf = {}
for i, r in enumerate(csv.DictReader(open(MASTER_FLAT, encoding="utf-8", errors="replace"))):
    if item_key(r.get("item_text_english"), r.get("item_text_original")):
        r["_mid"] = str(i); mf[str(i)] = r

disp = {d["row_id"]: d for d in csv.DictReader(open(os.path.join(OUT, "item_disposition.csv"), encoding="utf-8"))}

def domain_for(mid):
    d = disp.get(mid)
    if not d: return ""
    if d["disposition"] == "already_coded": return d["inherit_final_coding"]
    if d["disposition"] == "internal_dup":  return domain_for(d["maps_to_row_id"])
    # new:
    r = mf.get(mid)
    return coded_new.get(item_key(r.get("item_text_english"), r.get("item_text_original")), "")

out_rows, assigned = [], 0
for mid, r in mf.items():
    dom = domain_for(mid)
    if dom: assigned += 1
    rr = dict(r); rr["final_coding"] = dom; out_rows.append(rr)

cols = [c for c in out_rows[0].keys() if c != "_mid"] + ["final_coding"]
with open(os.path.join(OUT, "master_flat_with_rsf.csv"), "w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore"); w.writeheader(); w.writerows(out_rows)
print(f"item rows total: {len(out_rows)} | assigned an RSF domain: {assigned} | still blank: {len(out_rows)-assigned}")
print("-> output/master_flat_with_rsf.csv")
