#!/usr/bin/env python3
"""
STEP 1 — item-level EXACT de-duplication of the Phase-2 master_flat.

Removes only items whose text is an EXACT match (case/whitespace-insensitive) to:
  (a) an item already coded in the Phase-1 32,444 set  -> so Item Coding never re-codes it;
  (b) another Phase-2 item already kept                 -> internal duplicate, keep the first.
Everything else is a NEW item to be coded. No scale-name collapsing (that would drop
distinct same-named-scale variants). Fully reversible.

Reads : master_flat.csv , CODED_32444 (confirmatory_Step3_finished.xlsx: Item, final_coding)
Writes: output/01_new_items.csv        NEW items only (full master_flat columns + _mid)
        output/item_disposition.csv     EVERY item row + disposition, for the step-4 re-map:
                                         disposition = new | already_coded | internal_dup
                                         inherit_final_coding  (for already_coded)
                                         maps_to_row_id        (for internal_dup)
"""
import csv, os, sys
from openpyxl import load_workbook
sys.path.insert(0, os.path.dirname(__file__))
from common_paths import MASTER_FLAT, CODED_32444, STEP4_FINAL, OUT, norm, item_key
csv.field_size_limit(2**30)

# --- already-coded reference: exact item text -> final_coding (RSF domain) ---
# Keyed on BOTH English and original text, so a non-English Phase-2 item (which has no
# English translation) still matches an already-coded item via its ORIGINAL text.
coded = {}
# (1) English 'Item' -> final_coding, from the coded confirmatory file
eng2fc = {}
wb = load_workbook(CODED_32444, read_only=True); ws = wb.active
h = [c.value for c in next(ws.iter_rows(max_row=1))]
ii, fc = h.index("Item"), h.index("final_coding")
for row in ws.iter_rows(min_row=2, values_only=True):
    k = norm(row[ii])
    if k: eng2fc.setdefault(k, row[fc]); coded.setdefault(k, row[fc])
wb.close()
# (2) add ORIGINAL texts from step4_final, mapping each to its English item's final_coding
wb = load_workbook(STEP4_FINAL, read_only=True); ws = wb.active
h = [c.value for c in next(ws.iter_rows(max_row=1))]
oi, ei = h.index("item_text_original"), h.index("item_text_english")
for row in ws.iter_rows(min_row=2, values_only=True):
    e = norm(row[ei]); o = norm(row[oi]); f = eng2fc.get(e, "")
    if o and o not in coded: coded[o] = f
    if e and e not in coded: coded[e] = f
wb.close()
print(f"already-coded reference (32,444, original+english keys): {len(coded)}")

# --- Phase-2 master_flat item rows ---
rows = []
with open(MASTER_FLAT, encoding="utf-8", errors="replace") as f:
    for i, r in enumerate(csv.DictReader(f)):
        if item_key(r.get("item_text_english"), r.get("item_text_original")):
            r["_mid"] = str(i)
            rows.append(r)
print(f"master_flat item-bearing rows: {len(rows)}")

new_items, disp, seen = [], [], {}
for r in rows:
    k = item_key(r.get("item_text_english"), r.get("item_text_original"))
    if k in coded:
        disp.append({"row_id": r["_mid"], "disposition": "already_coded",
                     "inherit_final_coding": coded[k], "maps_to_row_id": ""})
    elif k in seen:
        disp.append({"row_id": r["_mid"], "disposition": "internal_dup",
                     "inherit_final_coding": "", "maps_to_row_id": seen[k]})
    else:
        seen[k] = r["_mid"]
        new_items.append(r)
        disp.append({"row_id": r["_mid"], "disposition": "new",
                     "inherit_final_coding": "", "maps_to_row_id": ""})

os.makedirs(OUT, exist_ok=True)
cols = [c for c in rows[0].keys()]
with open(os.path.join(OUT, "01_new_items.csv"), "w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore"); w.writeheader(); w.writerows(new_items)
with open(os.path.join(OUT, "item_disposition.csv"), "w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=["row_id","disposition","inherit_final_coding","maps_to_row_id"])
    w.writeheader(); w.writerows(disp)

from collections import Counter
c = Counter(d["disposition"] for d in disp)
print(f"  new (to code)        : {c['new']}")
print(f"  already_coded (skip) : {c['already_coded']}  (inherit RSF from 32,444)")
print(f"  internal_dup (skip)  : {c['internal_dup']}  (inherit from kept twin)")
print(f"  total item rows      : {len(rows)}  (all accounted for -> re-mappable)")
print("-> output/01_new_items.csv , output/item_disposition.csv")
